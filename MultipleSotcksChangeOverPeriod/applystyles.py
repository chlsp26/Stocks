from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import date
from string import ascii_uppercase
import itertools

today = date.today()
path = "D:\\Stocks\\" + str(today) + "\\stocks_{0}.xlsx".format(today)
wb = load_workbook(path)
ws = wb.active
ft = Font(color="FFFFFF", bold=True)
for columns in range(4,len(ws[1])+1,1):
    for rows in range(2,len(ws['D'])+1,1): 
        if not(ws.cell(row=rows,column=columns).value is None) and ws.cell(row=rows,column=columns).value < 0:
            cell = ws.cell(row=rows,column=columns)
            cell.fill = PatternFill("solid", fgColor="FF0000")
            cell.font = ft
        elif not(ws.cell(row=rows,column=columns).value is None) and ws.cell(row=rows,column=columns).value >= 0:
            cell = ws.cell(row=rows,column=columns)
            cell.fill = PatternFill("solid", fgColor="008000")
            cell.font = ft

def iter_all_strings():
    for size in itertools.count(1):
        for s in itertools.product(ascii_uppercase, repeat=size):
            yield "".join(s)

i = 1
for s in iter_all_strings():
    ws.column_dimensions[s].width = 10.5
    i += 1
    if i == len(ws[1]) + 1:
        break

path = "D:\\Stocks\\" + str(today) + "\\finalstocks_{0}.xlsx".format(today)
wb.save(path)